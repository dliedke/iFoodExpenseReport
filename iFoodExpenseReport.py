import os
import re
import json
import time
from datetime import datetime, timedelta

from selenium import webdriver
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

DIAS = 30  # janela de pedidos a considerar
MAX_PEDIDOS = 62  # teto de links coletados (evita varrer pedidos antigos demais)
CACHE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cache_pedidos.json")


def guid_do_href(href):
    """https://www.ifood.com.br/pedido/<uuid> -> '<uuid>'"""
    m = re.search(r"/pedido/([0-9a-fA-F-]{8,})", href)
    return m.group(1) if m else href


def carregar_cache():
    """Cache local em disco: guid -> {restaurante, total, data(iso)}."""
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def salvar_cache(cache):
    with open(CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


def parse_total(texto):
    """'R$ 66,10' -> 66.10"""
    m = re.search(r"([\d.]+,\d{2})", texto.replace("\xa0", " "))
    if not m:
        return None
    return float(m.group(1).replace(".", "").replace(",", "."))


def parse_data(texto):
    """'03/06/2026 • 18:51' -> datetime"""
    texto = texto.replace("\xa0", " ").strip()
    m = re.search(r"(\d{2}/\d{2}/\d{4})", texto)
    if not m:
        return None
    return datetime.strptime(m.group(1), "%d/%m/%Y")


def extrair_restaurante(driver, page_text):
    """Nome do restaurante: tenta seletores comuns, cai pro texto."""
    for sel in ["h2.order-details-header__title", ".order-details-header__title",
                "h1", "[class*='restaurant']", "[class*='merchant']",
                "[data-test-id*='merchant']", "h2"]:
        try:
            txt = driver.find_element(By.CSS_SELECTOR, sel).text.strip()
            if txt:
                return txt
        except Exception:
            continue
    # fallback: primeira linha não vazia do texto
    for linha in page_text.splitlines():
        linha = linha.strip()
        if linha:
            return linha
    return ""


def extrair_data(driver, page_text):
    """Data do pedido: seletor específico, cai pro 1º dd/mm/aaaa do texto."""
    for sel in ["p.order-details-footer-info",
                "[data-test-id*='date']", "[class*='date']"]:
        try:
            txt = driver.find_element(By.CSS_SELECTOR, sel).text
            d = parse_data(txt)
            if d:
                return d
        except Exception:
            continue
    return parse_data(page_text)  # regex acha o 1º dd/mm/aaaa na página


def extrair_total(driver, page_text):
    """Valor total: seletor específico, cai pro valor após a palavra 'Total'."""
    # seletor exato: o <span> do valor dentro do <p data-test-id="order-details-value">
    # (pula o primeiro span, que é o label "Total")
    for sel in ["p[data-test-id='order-details-value'] span:not(.order-details-value__label)",
                "p[data-test-id='order-details-value']",
                "[data-test-id*='order-details-value']",
                "[data-test-id*='total']", "[class*='total']"]:
        try:
            txt = driver.find_element(By.CSS_SELECTOR, sel).text
            v = parse_total(txt)
            if v is not None:
                return v
        except Exception:
            continue
    # fallback: pega o valor logo após "Total" (evita Subtotal usando \bTotal)
    txt = page_text.replace("\xa0", " ")
    m = re.search(r"(?<![Ss]ub)\bTotal\b[^\d]*([\d.]+,\d{2})", txt, re.IGNORECASE)
    if m:
        return float(m.group(1).replace(".", "").replace(",", "."))
    # último recurso: maior valor R$ da página (geralmente é o total)
    valores = [float(x.replace(".", "").replace(",", "."))
               for x in re.findall(r"R\$\s*([\d.]+,\d{2})", txt)]
    return max(valores) if valores else None


def build_driver():
    """Edge anti-detecção + perfil persistente (login e clearance do Cloudflare
    sobrevivem entre execuções)."""
    profile_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "edge_profile")

    options = EdgeOptions()
    options.add_argument(f"--user-data-dir={profile_dir}")
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--lang=pt-BR")
    # remove os sinais de automação que o Cloudflare procura
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)

    # Selenium Manager baixa o msedgedriver automaticamente
    driver = webdriver.Edge(options=options)

    # esconde navigator.webdriver antes de qualquer script da página rodar
    driver.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {"source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"},
    )
    return driver


def main():
    driver = build_driver()
    wait = WebDriverWait(driver, 20)

    driver.get("https://www.ifood.com.br/pedidos")

    # --- Espera o usuário autenticar manualmente ---
    input("Faça login no iFood no navegador aberto e pressione ENTER aqui para continuar...")

    # Garante que estamos na página de pedidos
    driver.get("https://www.ifood.com.br/pedidos")
    time.sleep(3)

    # --- Coleta os links de pedidos ---
    # Lazy load: faz scroll (e clica em "Ver mais", se houver) até parar de
    # aparecer pedido novo. Conta links, não altura da página (mais confiável).
    def coletar_links():
        anchors = driver.find_elements(By.CSS_SELECTOR, "a[href*='/pedido/']")
        achados = []
        vistos = set()
        for a in anchors:
            href = a.get_attribute("href")
            if href and "/pedido/" in href and href not in vistos:
                vistos.add(href)
                achados.append(href)
        return achados

    def clicar_ver_mais():
        # botão "Ver mais pedidos" da lista (seletor exato + fallbacks por texto)
        seletores = [
            (By.CSS_SELECTOR, "button[aria-label='Ver mais pedidos']"),
            (By.XPATH, "//button[contains(translate(., 'VERMAISPDOSCARG', 'vermaispdoscarg'), 'ver mais')]"),
            (By.XPATH, "//button[contains(translate(., 'CARGEMOIS', 'cargemois'), 'carregar mais')]"),
        ]
        for how, what in seletores:
            try:
                btn = driver.find_element(how, what)
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
                time.sleep(0.5)
                driver.execute_script("arguments[0].click();", btn)
                return True
            except Exception:
                continue
        return False

    links = []
    estagnado = 0
    for tentativa in range(60):  # teto alto; sai cedo se estabilizar
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)
        clicar_ver_mais()
        time.sleep(1)

        atual = coletar_links()
        if len(atual) > len(links):
            links = atual
            estagnado = 0
            print(f"  ...carregados {len(links)} pedidos")
        else:
            estagnado += 1
            if estagnado >= 4:  # 4 rodadas seguidas sem novidade -> fim
                break

        if len(links) >= MAX_PEDIDOS:  # já temos o suficiente -> para de rolar
            break

    links = coletar_links()[:MAX_PEDIDOS]

    print(f"Encontrados {len(links)} pedidos. Coletando detalhes...")

    limite = datetime.now() - timedelta(days=DIAS)
    cache = carregar_cache()
    print(f"Cache: {len(cache)} pedidos em disco.")
    pedidos = []
    fora_seguidos = 0  # pedidos vêm do mais novo p/ mais antigo: para ao sair da janela

    for idx, href in enumerate(links, 1):
        guid = guid_do_href(href)
        print(f"\n[{idx}/{len(links)}] {guid}")
        try:
            if guid in cache:
                c = cache[guid]
                restaurante = c["restaurante"]
                total = c["total"]
                data_dt = datetime.fromisoformat(c["data"]) if c.get("data") else None
                print("    [cache] usando dados salvos")
            else:
                driver.get(href)
                time.sleep(3)  # deixa o conteúdo carregar (SPA)
                page_text = driver.find_element(By.TAG_NAME, "body").text

                restaurante = extrair_restaurante(driver, page_text)
                data_dt = extrair_data(driver, page_text)
                total = extrair_total(driver, page_text)

                cache[guid] = {
                    "restaurante": restaurante,
                    "total": total,
                    "data": data_dt.isoformat() if data_dt else None,
                }
                salvar_cache(cache)  # persiste a cada pedido novo

            print(f"    restaurante : {restaurante!r}")
            print(f"    data        : {data_dt.strftime('%d/%m/%Y') if data_dt else None}")
            print(f"    total       : {total}")

            if data_dt and data_dt < limite:
                fora_seguidos += 1
                print(f"    -> fora da janela de {DIAS} dias, ignorado")
                if fora_seguidos >= 2:  # 2 seguidos antigos -> resto é mais antigo ainda
                    print("    -> janela de 30 dias ultrapassada, parando coleta")
                    break
                continue

            fora_seguidos = 0
            pedidos.append({
                "restaurante": restaurante,
                "total": total,
                "data": data_dt,
            })

        except Exception as e:
            print(f"    ERRO: {e}")
            continue

    try:
        driver.quit()
    except Exception:
        pass

    # --- Monta o XLSX ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos 30 dias"

    header_fill = PatternFill("solid", fgColor="EA1D2C")  # vermelho iFood
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(*[Side(style="thin", color="DDDDDD")] * 4)

    headers = ["Restaurante", "Total (R$)", "Data"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    pedidos.sort(key=lambda p: p["data"] or datetime.min, reverse=True)

    row = 2
    for p in pedidos:
        ws.cell(row=row, column=1, value=p["restaurante"]).border = border
        c2 = ws.cell(row=row, column=2, value=p["total"])
        c2.number_format = "#,##0.00"
        c2.border = border
        c3 = ws.cell(row=row, column=3, value=p["data"])
        c3.number_format = "DD/MM/YYYY"
        c3.border = border
        row += 1

    valores = [p["total"] for p in pedidos if p["total"] is not None]
    total_geral = sum(valores)
    media = total_geral / len(valores) if valores else 0

    row += 1
    ws.cell(row=row, column=1, value="Total geral").font = Font(bold=True)
    cg = ws.cell(row=row, column=2, value=total_geral)
    cg.number_format = "#,##0.00"
    cg.font = Font(bold=True)

    row += 1
    ws.cell(row=row, column=1, value="Média").font = Font(bold=True)
    cm = ws.cell(row=row, column=2, value=media)
    cm.number_format = "#,##0.00"
    cm.font = Font(bold=True)

    row += 1
    ws.cell(row=row, column=1, value="Qtd. pedidos").font = Font(bold=True)
    ws.cell(row=row, column=2, value=len(valores)).font = Font(bold=True)

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15

    nome_arquivo = f"pedidos_ifood_{DIAS}dias.xlsx"
    wb.save(nome_arquivo)
    print(f"\nArquivo salvo: {nome_arquivo}")
    print(f"Total geral: R$ {total_geral:.2f} | Média: R$ {media:.2f} | {len(valores)} pedidos")

    # Abre a planilha automaticamente (Windows)
    try:
        os.startfile(os.path.abspath(nome_arquivo))
    except Exception as e:
        print(f"Nao consegui abrir a planilha automaticamente: {e}")


if __name__ == "__main__":
    main()